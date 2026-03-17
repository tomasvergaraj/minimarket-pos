from io import BytesIO
from datetime import datetime, date
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session

from app.models.sale import Sale, SaleItem, SaleStatus
from app.models.product import Product

CHILE_TZ = ZoneInfo("America/Santiago")

PAYMENT_LABELS = {
    "cash": "Efectivo",
    "card": "Tarjeta",
    "mixed": "Mixto",
    "transfer": "Transferencia",
}


def _to_chile_time(dt: datetime) -> datetime:
    """Convert a naive UTC datetime to Chile local time."""
    if dt is None:
        return dt
    return dt.replace(tzinfo=ZoneInfo("UTC")).astimezone(CHILE_TZ)


def _fmt_dt(dt: datetime) -> str:
    local = _to_chile_time(dt)
    return local.strftime("%Y-%m-%d %H:%M")


def generate_sales_report(db: Session, date_from: date, date_to: date) -> BytesIO:
    wb = Workbook()

    # ── Sheet 1: Summary by sale ──────────────────────────────────────────────
    ws_sales = wb.active
    ws_sales.title = "Ventas"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    sale_headers = ["N° Venta", "Fecha (Chile)", "Método Pago", "Neto", "IVA", "Total", "Estado", "Productos"]
    ws_sales.append(sale_headers)
    for cell in ws_sales[1]:
        cell.font = header_font
        cell.fill = header_fill

    sales = (
        db.query(Sale)
        .filter(Sale.created_at >= datetime.combine(date_from, datetime.min.time()))
        .filter(Sale.created_at <= datetime.combine(date_to, datetime.max.time()))
        .order_by(Sale.sale_number)
        .all()
    )

    for sale in sales:
        products_str = "; ".join(f"{i.product_name} x{i.quantity}" for i in sale.items)
        net_subtotal = round(float(sale.total) - float(sale.tax_amount), 2)
        payment_label = PAYMENT_LABELS.get(
            sale.payment_method.value if hasattr(sale.payment_method, "value") else str(sale.payment_method),
            str(sale.payment_method),
        )
        ws_sales.append([
            sale.sale_number,
            _fmt_dt(sale.created_at),
            payment_label,
            net_subtotal,
            float(sale.tax_amount),
            float(sale.total),
            "Completada" if sale.status == SaleStatus.COMPLETED else "Anulada",
            products_str,
        ])

    # Summary row
    completed = [s for s in sales if s.status == SaleStatus.COMPLETED]
    ws_sales.append([])
    ws_sales.append(["TOTAL COMPLETADAS", "", "", "", "", sum(float(s.total) for s in completed)])

    for col in ws_sales.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_sales.column_dimensions[col[0].column_letter].width = min(max_length + 2, 45)

    # ── Sheet 2: Product breakdown (ventas x cantidad) ────────────────────────
    ws_prod = wb.create_sheet(title="Productos Vendidos")

    prod_headers = ["SKU", "Producto", "Cantidad Total", "Ventas (boletas)", "Precio Unit. Prom.", "Total Vendido"]
    ws_prod.append(prod_headers)
    for cell in ws_prod[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Aggregate by product across all completed sales in range
    product_totals: dict[str, dict] = {}
    for sale in completed:
        for item in sale.items:
            key = item.product_sku or item.product_name
            if key not in product_totals:
                product_totals[key] = {
                    "sku": item.product_sku,
                    "name": item.product_name,
                    "qty": 0,
                    "sale_count": 0,
                    "revenue": 0.0,
                    "prices": [],
                }
            product_totals[key]["qty"] += item.quantity
            product_totals[key]["sale_count"] += 1
            product_totals[key]["revenue"] += float(item.subtotal)
            product_totals[key]["prices"].append(float(item.unit_price))

    for key in sorted(product_totals, key=lambda k: product_totals[k]["revenue"], reverse=True):
        pt = product_totals[key]
        avg_price = round(sum(pt["prices"]) / len(pt["prices"]), 2) if pt["prices"] else 0
        ws_prod.append([
            pt["sku"],
            pt["name"],
            pt["qty"],
            pt["sale_count"],
            avg_price,
            round(pt["revenue"], 2),
        ])

    # Grand total row
    ws_prod.append([])
    ws_prod.append(["TOTAL", "", sum(pt["qty"] for pt in product_totals.values()), "", "", round(sum(pt["revenue"] for pt in product_totals.values()), 2)])

    for col in ws_prod.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws_prod.column_dimensions[col[0].column_letter].width = min(max_length + 2, 45)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_inventory_report(db: Session) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    headers = ["SKU", "Código Barra", "Nombre", "Categoría", "Stock", "Stock Mín.", "P. Costo", "P. Venta", "Activo"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    products = db.query(Product).order_by(Product.name).all()

    for p in products:
        ws.append([
            p.sku,
            p.barcode or "",
            p.name,
            p.category or "",
            p.stock,
            p.min_stock,
            float(p.cost_price),
            float(p.sell_price),
            "Sí" if p.is_active else "No",
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
